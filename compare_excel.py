#!/usr/bin/env python3
"""
Compare two Excel workbooks cell by cell.

Usage:
  python3 compare_excel.py old.xlsx new.xlsx
  python3 compare_excel.py old.xlsx new.xlsx --html diff_report.html
  python3 compare_excel.py old.xlsx new.xlsx --data-only

Notes:
  - Supports .xlsx/.xlsm files through openpyxl.
  - By default it compares formulas as formulas. Use --data-only to compare
    cached calculated values saved in the workbook.
"""

from __future__ import annotations

import argparse
import csv
import html
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ModuleNotFoundError:
    print(
        "Missing dependency: openpyxl\n"
        "Install it with:\n"
        "  python3 -m pip install openpyxl",
        file=sys.stderr,
    )
    raise SystemExit(2)


@dataclass(frozen=True)
class Difference:
    sheet: str
    cell: str
    left_value: Any
    right_value: Any


def normalize(value: Any, trim: bool, empty_equals_blank: bool) -> Any:
    if trim and isinstance(value, str):
        value = value.strip()

    if empty_equals_blank and value == "":
        return None

    return value


def sheet_names(left_wb: Any, right_wb: Any) -> list[str]:
    names = sorted(set(left_wb.sheetnames) | set(right_wb.sheetnames))
    return names


def max_dimensions(left_sheet: Any | None, right_sheet: Any | None) -> tuple[int, int]:
    max_row = max(
        left_sheet.max_row if left_sheet is not None else 0,
        right_sheet.max_row if right_sheet is not None else 0,
    )
    max_col = max(
        left_sheet.max_column if left_sheet is not None else 0,
        right_sheet.max_column if right_sheet is not None else 0,
    )
    return max_row, max_col


def cell_value(sheet: Any | None, row: int, col: int) -> Any:
    if sheet is None:
        return None

    if row > sheet.max_row or col > sheet.max_column:
        return None

    return sheet.cell(row=row, column=col).value


def compare_workbooks(
    left_path: Path,
    right_path: Path,
    *,
    data_only: bool,
    trim: bool,
    empty_equals_blank: bool,
) -> list[Difference]:
    left_wb = load_workbook(left_path, data_only=data_only, read_only=True)
    right_wb = load_workbook(right_path, data_only=data_only, read_only=True)

    differences: list[Difference] = []

    for sheet_name in sheet_names(left_wb, right_wb):
        left_sheet = left_wb[sheet_name] if sheet_name in left_wb.sheetnames else None
        right_sheet = right_wb[sheet_name] if sheet_name in right_wb.sheetnames else None
        max_row, max_col = max_dimensions(left_sheet, right_sheet)

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                left_value = normalize(
                    cell_value(left_sheet, row, col), trim, empty_equals_blank
                )
                right_value = normalize(
                    cell_value(right_sheet, row, col), trim, empty_equals_blank
                )

                if left_value != right_value:
                    cell = f"{get_column_letter(col)}{row}"
                    differences.append(
                        Difference(sheet_name, cell, left_value, right_value)
                    )

    left_wb.close()
    right_wb.close()
    return differences


def write_csv(path: Path, differences: Iterable[Difference]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(["sheet", "cell", "file1_value", "file2_value"])
        for diff in differences:
            writer.writerow([diff.sheet, diff.cell, diff.left_value, diff.right_value])


def display_value(value: Any) -> str:
    if value is None:
        return "<blank>"
    return str(value)


def write_html(path: Path, file1: Path, file2: Path, differences: list[Difference]) -> None:
    sheet_counts: dict[str, int] = {}
    for diff in differences:
        sheet_counts[diff.sheet] = sheet_counts.get(diff.sheet, 0) + 1

    summary_rows = "\n".join(
        "<tr>"
        f"<td>{html.escape(sheet)}</td>"
        f"<td>{count}</td>"
        "</tr>"
        for sheet, count in sorted(sheet_counts.items())
    )

    grouped_rows: list[str] = []
    current_sheet = None
    for diff in differences:
        if diff.sheet != current_sheet:
            current_sheet = diff.sheet
            grouped_rows.append(
                f"<tr class=\"sheet-row\"><th colspan=\"4\">{html.escape(diff.sheet)}</th></tr>"
            )

        grouped_rows.append(
            "<tr>"
            f"<td class=\"cell-ref\">{html.escape(diff.cell)}</td>"
            f"<td class=\"old-value\">{html.escape(display_value(diff.left_value))}</td>"
            f"<td class=\"new-value\">{html.escape(display_value(diff.right_value))}</td>"
            f"<td>{html.escape(diff.sheet)}!{html.escape(diff.cell)}</td>"
            "</tr>"
        )

    diff_rows = "\n".join(grouped_rows)
    empty_state = (
        "<section class=\"empty\">No cell differences found.</section>"
        if not differences
        else ""
    )

    page = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Excel Diff Report</title>
  <style>
    :root {{
      color-scheme: light;
      --bg: #f6f7f9;
      --panel: #ffffff;
      --text: #17202a;
      --muted: #64748b;
      --line: #d9dee8;
      --old: #fff1f2;
      --old-border: #f0a8b5;
      --new: #eefaf3;
      --new-border: #91d5a8;
      --accent: #2457a6;
    }}
    * {{ box-sizing: border-box; }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--text);
      font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Arial, sans-serif;
      line-height: 1.45;
    }}
    header {{
      background: var(--panel);
      border-bottom: 1px solid var(--line);
      padding: 24px 32px;
    }}
    h1 {{
      margin: 0 0 8px;
      font-size: 28px;
      letter-spacing: 0;
    }}
    .files {{
      display: grid;
      gap: 4px;
      color: var(--muted);
      font-size: 14px;
      word-break: break-all;
    }}
    main {{
      max-width: 1200px;
      margin: 0 auto;
      padding: 24px;
    }}
    .stats {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
      gap: 12px;
      margin-bottom: 18px;
    }}
    .stat {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      padding: 14px 16px;
    }}
    .stat strong {{
      display: block;
      font-size: 24px;
      color: var(--accent);
    }}
    .stat span {{
      color: var(--muted);
      font-size: 13px;
    }}
    section {{
      background: var(--panel);
      border: 1px solid var(--line);
      border-radius: 8px;
      overflow: hidden;
      margin-bottom: 18px;
    }}
    h2 {{
      margin: 0;
      padding: 14px 16px;
      font-size: 16px;
      border-bottom: 1px solid var(--line);
    }}
    .table-wrap {{
      overflow: auto;
    }}
    table {{
      width: 100%;
      border-collapse: collapse;
      font-size: 14px;
    }}
    th, td {{
      border-bottom: 1px solid var(--line);
      padding: 9px 10px;
      text-align: left;
      vertical-align: top;
    }}
    thead th {{
      position: sticky;
      top: 0;
      background: #eef2f7;
      z-index: 1;
    }}
    .sheet-row th {{
      background: #dfe8f5;
      color: #173b70;
      font-size: 15px;
    }}
    .cell-ref {{
      width: 90px;
      font-weight: 700;
      color: var(--accent);
      white-space: nowrap;
    }}
    .old-value {{
      background: var(--old);
      border-left: 4px solid var(--old-border);
      white-space: pre-wrap;
    }}
    .new-value {{
      background: var(--new);
      border-left: 4px solid var(--new-border);
      white-space: pre-wrap;
    }}
    .empty {{
      padding: 24px;
      text-align: center;
      color: var(--muted);
    }}
  </style>
</head>
<body>
  <header>
    <h1>Excel Diff Report</h1>
    <div class="files">
      <div><strong>File 1:</strong> {html.escape(str(file1))}</div>
      <div><strong>File 2:</strong> {html.escape(str(file2))}</div>
    </div>
  </header>
  <main>
    <div class="stats">
      <div class="stat"><strong>{len(differences)}</strong><span>Different cells</span></div>
      <div class="stat"><strong>{len(sheet_counts)}</strong><span>Sheets with differences</span></div>
    </div>
    {empty_state}
    <section>
      <h2>Summary by Sheet</h2>
      <div class="table-wrap">
        <table>
          <thead><tr><th>Sheet</th><th>Different Cells</th></tr></thead>
          <tbody>{summary_rows}</tbody>
        </table>
      </div>
    </section>
    <section>
      <h2>Cell Differences</h2>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>Cell</th>
              <th>File 1 Value</th>
              <th>File 2 Value</th>
              <th>Address</th>
            </tr>
          </thead>
          <tbody>{diff_rows}</tbody>
        </table>
      </div>
    </section>
  </main>
</body>
</html>
"""
    path.write_text(page, encoding="utf-8")


def print_differences(differences: list[Difference], limit: int) -> None:
    if not differences:
        print("No cell differences found.")
        return

    print(f"Found {len(differences)} different cell(s).")
    print("sheet\tcell\tfile1_value\tfile2_value")

    for diff in differences[:limit]:
        print(f"{diff.sheet}\t{diff.cell}\t{diff.left_value!r}\t{diff.right_value!r}")

    hidden = len(differences) - limit
    if hidden > 0:
        print(f"... {hidden} more difference(s). Use --output differences.csv to save all.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare two Excel files and list cells with different content."
    )
    parser.add_argument("file1", type=Path, help="First Excel file (.xlsx/.xlsm)")
    parser.add_argument("file2", type=Path, help="Second Excel file (.xlsx/.xlsm)")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        help="Optional CSV output path for all differences.",
    )
    parser.add_argument(
        "--html",
        type=Path,
        default=Path("excel_diff_report.html"),
        help="HTML report output path. Default: excel_diff_report.html",
    )
    parser.add_argument(
        "--no-html",
        action="store_true",
        help="Do not generate the visual HTML report.",
    )
    parser.add_argument(
        "--data-only",
        action="store_true",
        help="Compare cached calculated values instead of formulas.",
    )
    parser.add_argument(
        "--trim",
        action="store_true",
        help="Ignore leading/trailing spaces in text cells.",
    )
    parser.add_argument(
        "--strict-empty",
        action="store_true",
        help="Treat empty string and blank cell as different. Default treats them equal.",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=50,
        help="Maximum number of differences to print to terminal. Default: 50",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    if not args.file1.exists():
        print(f"File not found: {args.file1}", file=sys.stderr)
        return 1
    if not args.file2.exists():
        print(f"File not found: {args.file2}", file=sys.stderr)
        return 1

    differences = compare_workbooks(
        args.file1,
        args.file2,
        data_only=args.data_only,
        trim=args.trim,
        empty_equals_blank=not args.strict_empty,
    )

    print_differences(differences, max(args.limit, 0))

    if args.output:
        write_csv(args.output, differences)
        print(f"Saved CSV: {args.output}")

    if not args.no_html:
        write_html(args.html, args.file1, args.file2, differences)
        print(f"Saved HTML report: {args.html}")

    return 1 if differences else 0


if __name__ == "__main__":
    raise SystemExit(main())
